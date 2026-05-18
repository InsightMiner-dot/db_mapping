"""
SVP Comment Parser
------------------
Reads an unstructured 'Comments' column from one Excel file,
uses Azure OpenAI to extract and structure the data,
maps values using a reference/mapping Excel file,
and writes a clean structured Excel output.
"""

import os
import json
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv
from openai import AzureOpenAI

load_dotenv(override=True)

# ── Azure OpenAI Setup ───────────────────────────────────────────────────────
client = AzureOpenAI(
    api_key=os.getenv("AZURE_OPENAI_API_KEY"),
    api_version=os.getenv("AZURE_OPENAI_API_VERSION", "2024-02-15-preview"),
    azure_endpoint=os.getenv("AZURE_OPENAI_ENDPOINT"),
)
DEPLOYMENT = os.getenv("AZURE_OPENAI_DEPLOYMENT", "gpt-4o")


# ── Step 1: Ask user for file paths ─────────────────────────────────────────
def get_inputs():
    print("\n" + "=" * 55)
    print("  SVP Comment Parser — Structured DB Builder")
    print("=" * 55)

    comments_path = input("\n📂 Path to Comments Excel file   : ").strip().strip('"')
    mapping_path  = input("📂 Path to Mapping Excel file    : ").strip().strip('"')
    output_path   = input("💾 Output file path (.xlsx)       : ").strip().strip('"')

    # Load files and show available columns
    comments_df = pd.read_excel(comments_path)
    comments_df.columns = [c.strip() for c in comments_df.columns]
    print(f"\n   Columns in Comments file: {list(comments_df.columns)}")
    comment_col = input("   Which column has the comments? : ").strip()

    mapping_df = pd.read_excel(mapping_path)
    mapping_df.columns = [c.strip() for c in mapping_df.columns]
    print(f"\n   Columns in Mapping file: {list(mapping_df.columns)}")
    print("   (These will be used as the structured output columns)")

    return comments_df, mapping_df, comment_col, output_path


# ── Step 2: Build dynamic prompts from the actual mapping columns ─────────────
def build_prompts(mapping_df: pd.DataFrame):
    """Dynamically build LLM prompts based on whatever columns exist in the mapping file."""

    # Get unique values per mapping column (for LLM context)
    unique_vals = {
        col: sorted(mapping_df[col].dropna().unique().tolist())
        for col in mapping_df.columns
    }

    # --- Parse prompt: tells LLM how to split the raw comment ---
    parse_prompt = """You are a financial analyst. Parse the raw comment string into individual variance line items.

Each item in the comment typically contains:
- A description or category name
- A variance amount in millions (e.g. -0.5M, +1.2M, (0.3M))
- A reason or qualifier (often in parentheses, e.g. Timing, Saving)

Rules:
- Split on commas, semicolons, or new lines to find individual items
- Extract: description, variance_amount (number, null if missing), reason (text)
- Return ONLY a valid JSON array, no markdown, no extra text

Example output:
[
  {"description": "Legal", "variance_amount": -0.2, "reason": "Saving"},
  {"description": "IS IT", "variance_amount": -0.2, "reason": "Timing"}
]"""

    # --- Map prompt: tells LLM to assign values from actual mapping columns ---
    col_descriptions = "\n".join(
        f'- "{col}": choose from {vals[:10]}{"..." if len(vals) > 10 else ""}'
        for col, vals in unique_vals.items()
    )

    map_prompt = f"""You are a financial data mapper.

Given a variance item (description, amount, reason), assign the best matching value
for each column below from the allowed values. Return null if nothing fits.

Columns and allowed values:
{col_descriptions}

Return ONLY a valid JSON object with exactly these keys: {list(unique_vals.keys())}
No markdown, no extra text."""

    return parse_prompt, map_prompt, unique_vals


# ── Step 3: LLM calls ────────────────────────────────────────────────────────
def llm(system: str, user: str) -> str:
    response = client.chat.completions.create(
        model=DEPLOYMENT,
        messages=[{"role": "system", "content": system}, {"role": "user", "content": user}],
        temperature=0.0,
        max_tokens=1500,
    )
    return response.choices[0].message.content.strip()


def parse_comment(raw: str, parse_prompt: str) -> list[dict]:
    if not raw or pd.isna(raw):
        return []
    result = llm(parse_prompt, f"Parse this comment:\n{raw}")
    result = re.sub(r"```(?:json)?|```", "", result).strip()
    try:
        return json.loads(result)
    except Exception:
        return [{"description": str(raw), "variance_amount": None, "reason": str(raw)}]


def map_item(item: dict, map_prompt: str, unique_vals: dict) -> dict:
    user_msg = (
        f"Variance item:\n{json.dumps(item)}\n\n"
        f"Available values:\n{json.dumps({k: v[:15] for k, v in unique_vals.items()})}"
    )
    result = llm(map_prompt, user_msg)
    result = re.sub(r"```(?:json)?|```", "", result).strip()
    try:
        return json.loads(result)
    except Exception:
        return {col: None for col in unique_vals}


# ── Step 4: Process all rows ─────────────────────────────────────────────────
def process(comments_df, mapping_df, comment_col, output_path):
    parse_prompt, map_prompt, unique_vals = build_prompts(mapping_df)
    mapping_cols = list(mapping_df.columns)
    other_cols = [c for c in comments_df.columns if c != comment_col]

    output_rows = []
    total = len(comments_df)

    print(f"\n🔄 Processing {total} rows...\n")

    for i, row in comments_df.iterrows():
        raw = row.get(comment_col, "")
        print(f"  [{i+1}/{total}] {str(raw)[:60]}...")

        items = parse_comment(raw, parse_prompt)
        if not items:
            items = [{"description": None, "variance_amount": None, "reason": None}]

        for item in items:
            mapped = map_item(item, map_prompt, unique_vals)
            out = {c: row.get(c) for c in other_cols}           # carry original cols
            out["Original Comment"]    = raw
            out["Description"]         = item.get("description")
            out["Variance Amount (M)"] = item.get("variance_amount")
            out["Reason"]              = item.get("reason")
            for col in mapping_cols:
                out[col] = mapped.get(col)
            output_rows.append(out)

    out_df = pd.DataFrame(output_rows)

    # Column order: mapping cols | extracted cols | original carry-overs
    col_order = (
        mapping_cols
        + ["Description", "Variance Amount (M)", "Reason", "Original Comment"]
        + [c for c in other_cols if c not in mapping_cols]
    )
    col_order = [c for c in col_order if c in out_df.columns]
    out_df = out_df[col_order]

    write_excel(out_df, output_path, mapping_cols)
    print(f"\n✅ Done! {len(out_df)} rows written to:\n   {output_path}")
    return out_df


# ── Step 5: Write formatted Excel ────────────────────────────────────────────
def write_excel(df: pd.DataFrame, path: str, mapping_cols: list):
    wb = Workbook()
    ws = wb.active
    ws.title = "Structured Output"

    HDR_MAP  = PatternFill("solid", start_color="1F3864")  # dark blue  → mapping cols
    HDR_EXT  = PatternFill("solid", start_color="375623")  # dark green → extracted cols
    HDR_ORIG = PatternFill("solid", start_color="595959")  # grey       → original cols
    ALT_FILL = PatternFill("solid", start_color="EBF3FB")  # light blue → alt rows
    hdr_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    body_font = Font(name="Arial", size=10)
    thin = Border(*[Side(style="thin")] * 4)

    extracted = ["Description", "Variance Amount (M)", "Reason", "Original Comment"]

    for ci, col in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.font = hdr_font
        cell.border = thin
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if col in mapping_cols:
            cell.fill = HDR_MAP
        elif col in extracted:
            cell.fill = HDR_EXT
        else:
            cell.fill = HDR_ORIG

    for ri, row in enumerate(df.itertuples(index=False), 2):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = body_font
            cell.border = thin
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            if ri % 2 == 0:
                cell.fill = ALT_FILL
            col_name = df.columns[ci - 1]
            if col_name == "Variance Amount (M)" and isinstance(val, (int, float)):
                cell.number_format = '#,##0.00;(#,##0.00);"-"'
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(
                    name="Arial", size=10, bold=True,
                    color="C00000" if val < 0 else "375623"
                )

    for ci, col in enumerate(df.columns, 1):
        max_w = max(len(col), df[col].astype(str).str.len().max() if not df.empty else 10)
        ws.column_dimensions[get_column_letter(ci)].width = min(max_w + 3, 40)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 30

    os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
    wb.save(path)


# ── Main ─────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    comments_df, mapping_df, comment_col, output_path = get_inputs()
    process(comments_df, mapping_df, comment_col, output_path)
